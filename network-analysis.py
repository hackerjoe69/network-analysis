import nmap
import requests
import socket
import subprocess
import datetime
import os
import json
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from bs4 import BeautifulSoup
import concurrent.futures
from colorama import Fore, Style, init

# Initialize colorama
init(autoreset=True)

class VulnerabilityScanner:
    def __init__(self):
        self.report_data = {
            "metadata": {
                "scan_date": datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
                "scanner_version": "1.2"
            },
            "results": []
        }
        self.red = Font(color="FF0000", bold=True)
        self.orange = Font(color="FFA500", bold=True)
        self.green = Font(color="00AA00", bold=True)
        self.high_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
        self.medium_fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
        self.low_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")

    def scan_network(self, target_range):
        """Perform network discovery using nmap"""
        print(f"{Fore.CYAN}[*] Discovering network hosts...{Style.RESET_ALL}")
        nm = nmap.PortScanner()
        nm.scan(hosts=target_range, arguments='-sn')
        
        live_hosts = []
        for host in nm.all_hosts():
            if nm[host].state() == 'up':
                live_hosts.append(host)
                print(f"{Fore.GREEN}[+] Found live host: {host}{Style.RESET_ALL}")
        
        return live_hosts

    def port_scan(self, host, ports='1-1024'):
        """Perform comprehensive port scanning"""
        print(f"{Fore.CYAN}[*] Scanning ports on {host}...{Style.RESET_ALL}")
        nm = nmap.PortScanner()
        nm.scan(hosts=host, ports=ports, arguments='-sV --script vulners')
        
        open_ports = []
        for proto in nm[host].all_protocols():
            ports = nm[host][proto].keys()
            for port in ports:
                service = nm[host][proto][port]
                open_ports.append({
                    "port": port,
                    "protocol": proto,
                    "service": service['name'],
                    "version": service['version'],
                    "vulnerabilities": self.check_vulnerabilities(service['product'], service['version'])
                })
                print(f"{Fore.YELLOW}[!] Open port: {port}/{proto} - {service['name']} {service['version']}{Style.RESET_ALL}")
        
        return open_ports

    def check_vulnerabilities(self, software, version):
        """Check for known vulnerabilities"""
        vulns = []
        cve_url = f"https://cve.mitre.org/cgi-bin/cvekey.cgi?keyword={software}+{version}"
        
        try:
            response = requests.get(cve_url, timeout=10)
            soup = BeautifulSoup(response.text, 'html.parser')
            table = soup.find('div', {'id': 'TableWithRules'})
            
            if table:
                rows = table.find_all('tr')[1:]  # Skip header
                for row in rows:
                    cols = row.find_all('td')
                    if len(cols) >= 2:
                        cve_id = cols[0].text.strip()
                        description = cols[1].text.strip()
                        severity = self.estimate_severity(description)
                        vulns.append({
                            "cve_id": cve_id,
                            "description": description,
                            "severity": severity
                        })
        except Exception as e:
            print(f"{Fore.RED}[-] Vulnerability check error: {e}{Style.RESET_ALL}")
        
        return vulns

    def estimate_severity(self, description):
        """Estimate vulnerability severity based on keywords"""
        description = description.lower()
        if 'remote code execution' in description or 'privilege escalation' in description:
            return "High"
        elif 'denial of service' in description or 'information disclosure' in description:
            return "Medium"
        else:
            return "Low"

    def web_app_scan(self, url):
        """Scan web application for common vulnerabilities"""
        print(f"{Fore.CYAN}[*] Scanning web application: {url}{Style.RESET_ALL}")
        findings = []
        
        # Check for common vulnerabilities
        checks = [
            ("SQL Injection", self.check_sql_injection, url),
            ("XSS Vulnerability", self.check_xss, url),
            ("Outdated Software", self.check_outdated_software, url),
            ("Insecure Headers", self.check_security_headers, url)
        ]
        
        with concurrent.futures.ThreadPoolExecutor() as executor:
            future_to_check = {executor.submit(func, param): name for name, func, param in checks}
            for future in concurrent.futures.as_completed(future_to_check):
                check_name = future_to_check[future]
                try:
                    result = future.result()
                    if result:
                        findings.append({"check": check_name, "result": result})
                except Exception as e:
                    print(f"{Fore.RED}[-] {check_name} check failed: {e}{Style.RESET_ALL}")
        
        return findings

    def check_sql_injection(self, url):
        """Check for basic SQL injection vulnerabilities"""
        test_payloads = ["'", "\"", "1' OR '1'='1"]
        vulnerable = False
        
        for payload in test_payloads:
            test_url = f"{url}?id={payload}"
            try:
                response = requests.get(test_url, timeout=5)
                if "error in your SQL syntax" in response.text.lower():
                    vulnerable = True
                    break
            except:
                continue
        
        return "Potential SQLi vulnerability detected" if vulnerable else "No obvious SQLi detected"

    def check_xss(self, url):
        """Check for basic XSS vulnerabilities"""
        test_payload = "<script>alert('XSS')</script>"
        test_url = f"{url}?search={test_payload}"
        
        try:
            response = requests.get(test_url, timeout=5)
            if test_payload in response.text:
                return "Potential XSS vulnerability detected"
        except:
            pass
        
        return "No obvious XSS detected"

    def check_outdated_software(self, url):
        """Check for outdated web software"""
        try:
            response = requests.get(url, timeout=5)
            headers = response.headers
            
            if 'server' in headers:
                server = headers['server']
                if 'Apache' in server:
                    version = server.split('/')[1] if '/' in server else 'unknown'
                    if version != 'unknown' and version < '2.4.53':
                        return f"Outdated Apache version ({version}) with known vulnerabilities"
                
                if 'X-Powered-By' in headers:
                    powered_by = headers['X-Powered-By']
                    if 'PHP' in powered_by:
                        version = powered_by.split('/')[1] if '/' in powered_by else 'unknown'
                        if version != 'unknown' and version < '8.0':
                            return f"Outdated PHP version ({version}) with known vulnerabilities"
            
            return "No obviously outdated software detected"
        except:
            return "Could not check software versions"

    def check_security_headers(self, url):
        """Check for missing security headers"""
        required_headers = [
            'X-Content-Type-Options',
            'X-Frame-Options',
            'Content-Security-Policy',
            'Strict-Transport-Security'
        ]
        
        missing_headers = []
        
        try:
            response = requests.get(url, timeout=5)
            headers = response.headers
            
            for header in required_headers:
                if header not in headers:
                    missing_headers.append(header)
            
            if missing_headers:
                return f"Missing security headers: {', '.join(missing_headers)}"
            else:
                return "All recommended security headers present"
        except:
            return "Could not check security headers"

    def generate_report(self, format_type='excel'):
        """Generate comprehensive vulnerability report"""
        print(f"{Fore.CYAN}[*] Generating {format_type} report...{Style.RESET_ALL}")
        
        if format_type == 'excel':
            self._generate_excel_report()
        elif format_type == 'json':
            self._generate_json_report()
        else:
            self._generate_text_report()

    def _generate_excel_report(self):
        """Generate Excel report with color coding"""
        wb = Workbook()
        ws = wb.active
        ws.title = "Vulnerability Report"
        
        # Header row
        headers = ["Host", "Port/Service", "Vulnerability", "Severity", "Recommendation"]
        ws.append(headers)
        
        for cell in ws[1]:
            cell.font = Font(bold=True)
        
        # Add data
        for result in self.report_data['results']:
            host = result['host']
            
            if 'ports' in result:
                for port_data in result['ports']:
                    port = port_data['port']
                    service = port_data['service']
                    version = port_data['version']
                    
                    if port_data['vulnerabilities']:
                        for vuln in port_data['vulnerabilities']:
                            row = [
                                host,
                                f"{port}/{port_data['protocol']} - {service} {version}",
                                f"{vuln['cve_id']}: {vuln['description']}",
                                vuln['severity'],
                                self._get_recommendation(service, version, vuln['severity'])
                            ]
                            ws.append(row)
                    else:
                        row = [
                            host,
                            f"{port}/{port_data['protocol']} - {service} {version}",
                            "No known vulnerabilities detected",
                            "None",
                            "None"
                        ]
                        ws.append(row)
            
            if 'web_findings' in result:
                for finding in result['web_findings']:
                    severity = "High" if "SQLi" in finding['check'] or "XSS" in finding['check'] else "Medium"
                    row = [
                        host,
                        "Web Application",
                        f"{finding['check']}: {finding['result']}",
                        severity,
                        self._get_recommendation("web", finding['check'], severity)
                    ]
                    ws.append(row)
        
        # Apply formatting
        for row in ws.iter_rows(min_row=2):
            severity_cell = row[3]
            if severity_cell.value == "High":
                for cell in row:
                    cell.fill = self.high_fill
                    cell.font = self.red
            elif severity_cell.value == "Medium":
                for cell in row:
                    cell.fill = self.medium_fill
                    cell.font = self.orange
            elif severity_cell.value == "Low":
                for cell in row:
                    cell.fill = self.low_fill
                    cell.font = self.green
        
        # Auto-size columns
        for col in ws.columns:
            max_length = 0
            column = col[0].column_letter
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(cell.value)
                except:
                    pass
            adjusted_width = (max_length + 2) * 1.2
            ws.column_dimensions[column].width = adjusted_width
        
        # Save file
        timestamp = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f"vulnerability_report_{timestamp}.xlsx"
        wb.save(filename)
        print(f"{Fore.GREEN}[+] Excel report saved as {filename}{Style.RESET_ALL}")

    def _generate_json_report(self):
        """Generate JSON report"""
        timestamp = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f"vulnerability_report_{timestamp}.json"
        
        with open(filename, 'w') as f:
            json.dump(self.report_data, f, indent=4)
        
        print(f"{Fore.GREEN}[+] JSON report saved as {filename}{Style.RESET_ALL}")

    def _generate_text_report(self):
        """Generate text report"""
        timestamp = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f"vulnerability_report_{timestamp}.txt"
        
        with open(filename, 'w') as f:
            f.write(f"Vulnerability Scan Report\n")
            f.write(f"Date: {self.report_data['metadata']['scan_date']}\n")
            f.write(f"Scanner Version: {self.report_data['metadata']['scanner_version']}\n\n")
            
            for result in self.report_data['results']:
                f.write(f"\n=== Host: {result['host']} ===\n")
                
                if 'ports' in result:
                    f.write("\nOpen Ports and Services:\n")
                    for port_data in result['ports']:
                        f.write(f"- {port_data['port']}/{port_data['protocol']}: {port_data['service']} {port_data['version']}\n")
                        
                        if port_data['vulnerabilities']:
                            f.write("  Vulnerabilities:\n")
                            for vuln in port_data['vulnerabilities']:
                                f.write(f"  * {vuln['cve_id']} ({vuln['severity']}): {vuln['description']}\n")
                                f.write(f"    Recommendation: {self._get_recommendation(port_data['service'], port_data['version'], vuln['severity'])}\n")
                        else:
                            f.write("  No known vulnerabilities detected\n")
                
                if 'web_findings' in result:
                    f.write("\nWeb Application Findings:\n")
                    for finding in result['web_findings']:
                        severity = "High" if "SQLi" in finding['check'] or "XSS" in finding['check'] else "Medium"
                        f.write(f"- {finding['check']}: {finding['result']} ({severity})\n")
                        f.write(f"  Recommendation: {self._get_recommendation('web', finding['check'], severity)}\n")
        
        print(f"{Fore.GREEN}[+] Text report saved as {filename}{Style.RESET_ALL}")

    def _get_recommendation(self, service, version, severity):
        """Generate actionable recommendations"""
        recommendations = {
            "general": {
                "High": "Immediate action required. Patch or mitigate as soon as possible.",
                "Medium": "Schedule remediation within the next 30 days.",
                "Low": "Consider addressing during next maintenance window."
            },
            "web": {
                "SQL Injection": "Implement parameterized queries and input validation. Consider using a Web Application Firewall (WAF).",
                "XSS Vulnerability": "Implement output encoding and Content Security Policy (CSP).",
                "Outdated Software": f"Upgrade to the latest stable version of {service}.",
                "Insecure Headers": "Configure web server to include missing security headers."
            },
            "Apache": "Upgrade to the latest stable version (2.4.53 or newer).",
            "nginx": "Upgrade to the latest stable version and ensure security configurations are applied.",
            "Microsoft-IIS": "Apply latest security patches from Microsoft and harden configuration.",
            "OpenSSH": "Upgrade to latest version and disable deprecated algorithms.",
            "MySQL": "Upgrade to latest version and review user privileges.",
            "PostgreSQL": "Upgrade to latest version and apply security best practices.",
            "Samba": "Upgrade to latest version and ensure insecure configurations are disabled."
        }
        
        if service in recommendations:
            if isinstance(recommendations[service], dict):
                return recommendations[service].get(severity, recommendations['general'][severity])
            return recommendations[service]
        elif service == "web":
            return recommendations['web'].get(version, recommendations['general'][severity])
        else:
            return recommendations['general'][severity]

    def run_scan(self, target_range, web_apps=None):
        """Run complete vulnerability scan"""
        print(f"{Fore.BLUE}\n[=== Starting Vulnerability Scan ===]{Style.RESET_ALL}")
        
        # Network discovery
        live_hosts = self.scan_network(target_range)
        
        if not live_hosts:
            print(f"{Fore.RED}[-] No live hosts found in the target range{Style.RESET_ALL}")
            return
        
        # Port scanning and vulnerability checks
        with concurrent.futures.ThreadPoolExecutor(max_workers=5) as executor:
            future_to_host = {executor.submit(self.port_scan, host): host for host in live_hosts}
            for future in concurrent.futures.as_completed(future_to_host):
                host = future_to_host[future]
                try:
                    ports = future.result()
                    self.report_data['results'].append({
                        "host": host,
                        "ports": ports
                    })
                except Exception as e:
                    print(f"{Fore.RED}[-] Error scanning {host}: {e}{Style.RESET_ALL}")
        
        # Web application scanning
        if web_apps:
            with concurrent.futures.ThreadPoolExecutor(max_workers=3) as executor:
                future_to_url = {executor.submit(self.web_app_scan, url): url for url in web_apps}
                for future in concurrent.futures.as_completed(future_to_url):
                    url = future_to_host[future]
                    try:
                        findings = future.result()
                        host = url.split('/')[2]  # Extract domain/host from URL
                        
                        # Find or create host entry
                        host_entry = None
                        for entry in self.report_data['results']:
                            if entry['host'] == host:
                                host_entry = entry
                                break
                        
                        if not host_entry:
                            host_entry = {"host": host}
                            self.report_data['results'].append(host_entry)
                        
                        host_entry['web_findings'] = findings
                    except Exception as e:
                        print(f"{Fore.RED}[-] Error scanning {url}: {e}{Style.RESET_ALL}")
        
        print(f"{Fore.BLUE}\n[=== Scan Completed ===]{Style.RESET_ALL}")

if __name__ == "__main__":
    scanner = VulnerabilityScanner()
    
    # Configuration
    target_network = "192.168.1.0/24"  # Adjust to your network
    web_applications = [
        "http://example.com",           # Add your web applications
        "https://app.example.com"
    ]
    
    # Run scan
    scanner.run_scan(target_network, web_applications)
    
    # Generate reports
    scanner.generate_report('excel')
    scanner.generate_report('json')
    scanner.generate_report('text')